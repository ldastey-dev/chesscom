"""Unit tests for chesscom.export.excel (ExcelReportWriter + SheetConfig)."""

from __future__ import annotations

import os

import pandas as pd
import pytest
from openpyxl import load_workbook
from openpyxl.styles import Font

from chesscom.export.excel import ExcelReportWriter, SheetConfig

# ---------------------------------------------------------------------------
# Helpers / fixtures
# ---------------------------------------------------------------------------


def _simple_df() -> pd.DataFrame:
    return pd.DataFrame(
        {
            "Username": ["alice", "bob", "charlie"],
            "Rating": [1500, 1600, 1700],
        }
    )


def _writer(tmp_path, sheets, base_name="Test Report"):
    return ExcelReportWriter(
        output_dir=str(tmp_path / "output"),
        base_name=base_name,
        sheets=sheets,
    )


# ===========================================================================
# File creation
# ===========================================================================


class TestFileCreation:
    def test_write_creates_file(self, tmp_path):
        sheet = SheetConfig(name="Data", dataframe=_simple_df())
        path = _writer(tmp_path, [sheet]).write()
        assert os.path.isfile(path)

    def test_write_returns_correct_path(self, tmp_path):
        sheet = SheetConfig(name="Data", dataframe=_simple_df())
        path = _writer(tmp_path, [sheet], base_name="My Report").write()
        assert path.endswith("My Report.xlsx")

    def test_write_creates_output_dir_if_absent(self, tmp_path):
        output_dir = tmp_path / "deep" / "nested" / "dir"
        assert not output_dir.exists()
        sheet = SheetConfig(name="Data", dataframe=_simple_df())
        ExcelReportWriter(
            output_dir=str(output_dir), base_name="R", sheets=[sheet]
        ).write()
        assert output_dir.exists()

    def test_write_empty_dataframe_still_creates_file(self, tmp_path):
        sheet = SheetConfig(name="Empty", dataframe=pd.DataFrame())
        path = _writer(tmp_path, [sheet]).write()
        assert os.path.isfile(path)


# ===========================================================================
# Sheet naming
# ===========================================================================


class TestSheetNaming:
    def test_single_sheet_name(self, tmp_path):
        sheet = SheetConfig(name="Members", dataframe=_simple_df())
        path = _writer(tmp_path, [sheet]).write()
        wb = load_workbook(path)
        assert "Members" in wb.sheetnames

    def test_multiple_sheet_names(self, tmp_path):
        sheets = [
            SheetConfig(name="Alpha", dataframe=_simple_df()),
            SheetConfig(name="Beta", dataframe=_simple_df()),
        ]
        path = _writer(tmp_path, sheets).write()
        wb = load_workbook(path)
        assert wb.sheetnames == ["Alpha", "Beta"]

    def test_sheet_order_preserved(self, tmp_path):
        names = ["Third", "First", "Second"]
        sheets = [SheetConfig(name=n, dataframe=_simple_df()) for n in names]
        path = _writer(tmp_path, sheets).write()
        wb = load_workbook(path)
        assert wb.sheetnames == names


# ===========================================================================
# Data content
# ===========================================================================


class TestDataContent:
    def test_header_row_written(self, tmp_path):
        sheet = SheetConfig(name="Data", dataframe=_simple_df())
        path = _writer(tmp_path, [sheet]).write()
        ws = load_workbook(path)["Data"]
        assert ws.cell(row=1, column=1).value == "Username"
        assert ws.cell(row=1, column=2).value == "Rating"

    def test_data_rows_written(self, tmp_path):
        sheet = SheetConfig(name="Data", dataframe=_simple_df())
        path = _writer(tmp_path, [sheet]).write()
        ws = load_workbook(path)["Data"]
        assert ws.cell(row=2, column=1).value == "alice"
        assert ws.cell(row=4, column=1).value == "charlie"

    def test_multiple_sheets_contain_correct_data(self, tmp_path):
        df1 = pd.DataFrame({"A": [1, 2]})
        df2 = pd.DataFrame({"B": [3, 4]})
        sheets = [
            SheetConfig(name="Sheet1", dataframe=df1),
            SheetConfig(name="Sheet2", dataframe=df2),
        ]
        path = _writer(tmp_path, sheets).write()
        wb = load_workbook(path)
        assert wb["Sheet1"].cell(row=2, column=1).value == 1
        assert wb["Sheet2"].cell(row=2, column=1).value == 3


# ===========================================================================
# Hyperlink injection
# ===========================================================================


class TestHyperlinkInjection:
    _URL_TPL = "https://www.chess.com/member/{value}"

    def test_hyperlink_applied_to_correct_column(self, tmp_path):
        sheet = SheetConfig(
            name="Members",
            dataframe=_simple_df(),
            hyperlink_column="Username",
            hyperlink_url_template=self._URL_TPL,
        )
        path = _writer(tmp_path, [sheet]).write()
        ws = load_workbook(path)["Members"]
        # Column 1 is Username; row 2 is first data row (row 1 = header)
        assert ws.cell(row=2, column=1).hyperlink is not None

    def test_hyperlink_url_expanded_correctly(self, tmp_path):
        sheet = SheetConfig(
            name="Members",
            dataframe=_simple_df(),
            hyperlink_column="Username",
            hyperlink_url_template=self._URL_TPL,
        )
        path = _writer(tmp_path, [sheet]).write()
        ws = load_workbook(path)["Members"]
        link = ws.cell(row=2, column=1).hyperlink
        assert link.target == "https://www.chess.com/member/alice"

    def test_all_data_rows_have_hyperlinks(self, tmp_path):
        df = _simple_df()
        sheet = SheetConfig(
            name="Members",
            dataframe=df,
            hyperlink_column="Username",
            hyperlink_url_template=self._URL_TPL,
        )
        path = _writer(tmp_path, [sheet]).write()
        ws = load_workbook(path)["Members"]
        for row in range(2, len(df) + 2):
            assert ws.cell(row=row, column=1).hyperlink is not None

    def test_hyperlink_font_is_blue_and_underlined(self, tmp_path):
        sheet = SheetConfig(
            name="Members",
            dataframe=_simple_df(),
            hyperlink_column="Username",
            hyperlink_url_template=self._URL_TPL,
        )
        path = _writer(tmp_path, [sheet]).write()
        ws = load_workbook(path)["Members"]
        font: Font = ws.cell(row=2, column=1).font
        assert font.color.rgb == "FF0000FF"
        assert font.underline == "single"

    def test_hyperlink_on_non_first_column(self, tmp_path):
        df = pd.DataFrame({"Score": [10, 20], "Username": ["x", "y"]})
        sheet = SheetConfig(
            name="S",
            dataframe=df,
            hyperlink_column="Username",
            hyperlink_url_template=self._URL_TPL,
        )
        path = _writer(tmp_path, [sheet]).write()
        ws = load_workbook(path)["S"]
        # Username is column 2 here
        assert ws.cell(row=2, column=2).hyperlink is not None

    def test_no_hyperlink_config_leaves_cells_clean(self, tmp_path):
        sheet = SheetConfig(name="Data", dataframe=_simple_df())
        path = _writer(tmp_path, [sheet]).write()
        ws = load_workbook(path)["Data"]
        assert ws.cell(row=2, column=1).hyperlink is None

    def test_hyperlink_column_not_in_dataframe_is_silently_skipped(self, tmp_path):
        sheet = SheetConfig(
            name="Data",
            dataframe=_simple_df(),
            hyperlink_column="NonExistent",
            hyperlink_url_template=self._URL_TPL,
        )
        path = _writer(tmp_path, [sheet]).write()
        assert os.path.isfile(path)

    def test_validation_error_when_template_missing(self, tmp_path):
        sheet = SheetConfig(
            name="Data",
            dataframe=_simple_df(),
            hyperlink_column="Username",
            hyperlink_url_template=None,
        )
        with pytest.raises(ValueError, match="hyperlink_url_template is required"):
            _writer(tmp_path, [sheet]).write()


# ===========================================================================
# Unique filename generation
# ===========================================================================


class TestUniqueFilename:
    def test_first_write_has_no_suffix(self, tmp_path):
        sheet = SheetConfig(name="S", dataframe=_simple_df())
        path = _writer(tmp_path, [sheet], base_name="Report").write()
        assert os.path.basename(path) == "Report.xlsx"

    def test_second_write_gets_suffix_1(self, tmp_path):
        sheet = SheetConfig(name="S", dataframe=_simple_df())
        w = _writer(tmp_path, [sheet], base_name="Report")
        w.write()
        path2 = _writer(tmp_path, [sheet], base_name="Report").write()
        assert os.path.basename(path2) == "Report_1.xlsx"

    def test_third_write_gets_suffix_2(self, tmp_path):
        sheet = SheetConfig(name="S", dataframe=_simple_df())
        for _ in range(2):
            _writer(tmp_path, [sheet], base_name="Report").write()
        path3 = _writer(tmp_path, [sheet], base_name="Report").write()
        assert os.path.basename(path3) == "Report_2.xlsx"

    def test_different_base_names_do_not_interfere(self, tmp_path):
        sheet = SheetConfig(name="S", dataframe=_simple_df())
        p1 = _writer(tmp_path, [sheet], base_name="Alpha").write()
        p2 = _writer(tmp_path, [sheet], base_name="Beta").write()
        assert os.path.basename(p1) == "Alpha.xlsx"
        assert os.path.basename(p2) == "Beta.xlsx"
